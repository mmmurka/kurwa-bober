import openpyxl


def print_cell_names(file_path):
    # Загружаем существующую таблицу
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active  # Выбираем активный лист

    # Проходимся по всем ячейкам и выводим их координаты и значения
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            print(f"Ячейка {cell.coordinate} содержит значение: {cell.value}")


# Пример использования
print_cell_names("prod.xlsx")
